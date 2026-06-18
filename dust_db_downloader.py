import csv
import datetime as dt
import math
import shutil
import sys
import threading
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    openpyxl = None
    Workbook = None


APP_NAME = "DustDB"
APP_VERSION = "v1"
APP_TITLE = f"{APP_NAME} {APP_VERSION}"
MAINTAINER = "杨威，2023级硕士研究生"
MAINTAINER_EMAIL = "yangw1663@gmail.com"
BLANK_LABEL = "<空值>"


def normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def display_value(value):
    return value if value != "" else BLANK_LABEL


def stored_value(value):
    return "" if value == BLANK_LABEL else value


def safe_filename(name):
    bad_chars = '<>:"/\\|?*'
    result = "".join("_" if ch in bad_chars else ch for ch in name)
    return result.strip() or "export"


class MultiSelectBox(ttk.Frame):
    def __init__(self, master, title, height=12):
        super().__init__(master)
        self.title = title

        header = ttk.Frame(self)
        header.pack(fill="x")
        ttk.Label(header, text=title).pack(side="left")
        ttk.Button(header, text="全选", width=6, command=self.select_all).pack(side="right", padx=(4, 0))
        ttk.Button(header, text="清空", width=6, command=self.clear).pack(side="right")

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, pady=(4, 0))
        self.listbox = tk.Listbox(body, selectmode="extended", exportselection=False, height=height)
        scrollbar = ttk.Scrollbar(body, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=scrollbar.set)
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def set_values(self, values, select_all=False, selected_values=None):
        self.listbox.delete(0, tk.END)
        selected_values = set(selected_values or [])
        for value in values:
            self.listbox.insert(tk.END, value)
        for index, value in enumerate(values):
            if select_all or value in selected_values:
                self.listbox.selection_set(index)

    def get_selected(self):
        return [self.listbox.get(i) for i in self.listbox.curselection()]

    def select_all(self):
        self.listbox.selection_set(0, tk.END)

    def clear(self):
        self.listbox.selection_clear(0, tk.END)


class DustDatabaseDownloader(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} - 滞尘数据库下载工具")
        self.geometry("1220x800")
        self.minsize(1040, 680)

        self.spectra_dir_var = tk.StringVar(value="")
        self.summary_path_var = tk.StringVar(value="")
        self.output_dir_var = tk.StringVar(value="")
        self.filter_logic_var = tk.StringVar(value="AND")
        self.copy_spectra_var = tk.BooleanVar(value=True)
        self.save_csv_var = tk.BooleanVar(value=True)
        self.save_xlsx_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="请选择汇总表路径，然后加载数据库索引。")
        self.count_var = tk.StringVar(value="未加载")
        self.value_search_var = tk.StringVar(value="")
        self.spectrum_sample_var = tk.StringVar(value="")
        self.spectrum_min_band_var = tk.StringVar(value="")
        self.spectrum_max_band_var = tk.StringVar(value="")
        self.spectrum_smooth_var = tk.IntVar(value=1)
        self.spectrum_normalize_var = tk.BooleanVar(value=False)
        self.spectrum_stats_var = tk.StringVar(value="未绘制光谱")

        self.rows = []
        self.headers = []
        self.filter_fields = []
        self.field_values = {}
        self.active_filters = {}
        self.current_field = ""
        self.current_visible_values = []
        self.parameter_box = None
        self.spectrum_selector = None
        self.spectrum_canvas = None
        self.spectrum_meta_tree = None
        self.current_spectrum = None
        self.load_lock = threading.Lock()

        self._build_ui()

    def _build_ui(self):
        self._configure_style()
        self._build_menu()

        outer = ttk.Frame(self, padding=12)
        outer.pack(fill="both", expand=True)

        title_frame = ttk.Frame(outer)
        title_frame.pack(fill="x")
        ttk.Label(title_frame, text=APP_TITLE, style="Title.TLabel").pack(side="left")
        ttk.Label(
            title_frame,
            text=f"滞尘数据库检索与下载 | 第一维护者：{MAINTAINER} | {MAINTAINER_EMAIL}",
            style="Subtle.TLabel",
        ).pack(side="left", padx=(16, 0))

        path_frame = ttk.LabelFrame(outer, text="路径")
        path_frame.pack(fill="x", pady=(10, 0))
        self._path_row(path_frame, "汇总表路径", self.summary_path_var, self.browse_summary_path, 0)
        self._path_row(path_frame, "光谱文件夹", self.spectra_dir_var, self.browse_spectra_dir, 1)
        self._path_row(path_frame, "导出位置", self.output_dir_var, self.browse_output_dir, 2)
        path_frame.columnconfigure(1, weight=1)

        action_frame = ttk.Frame(path_frame)
        action_frame.grid(row=3, column=1, sticky="ew", pady=(8, 4))
        ttk.Button(action_frame, text="加载汇总表", command=self.load_summary_async).pack(side="left")
        ttk.Button(action_frame, text="预览匹配数量", command=self.preview_count).pack(side="left", padx=8)
        ttk.Button(action_frame, text="下载选中数据", command=self.export_async).pack(side="left")
        ttk.Label(action_frame, textvariable=self.count_var, style="Subtle.TLabel").pack(side="left", padx=16)

        option_frame = ttk.LabelFrame(outer, text="导出与筛选逻辑")
        option_frame.pack(fill="x", pady=(10, 0))
        ttk.Label(option_frame, text="多字段关系").pack(side="left", padx=(8, 6), pady=8)
        ttk.Radiobutton(option_frame, text="同时满足 AND", variable=self.filter_logic_var, value="AND").pack(side="left")
        ttk.Radiobutton(option_frame, text="满足任一 OR", variable=self.filter_logic_var, value="OR").pack(side="left", padx=(8, 16))
        ttk.Checkbutton(option_frame, text="同时下载所选数据的光谱 txt", variable=self.copy_spectra_var).pack(side="left", padx=8)
        ttk.Checkbutton(option_frame, text="另存 CSV", variable=self.save_csv_var).pack(side="left", padx=8)
        ttk.Checkbutton(option_frame, text="另存 XLSX", variable=self.save_xlsx_var).pack(side="left", padx=8)

        self.notebook = ttk.Notebook(outer)
        self.notebook.pack(fill="both", expand=True, pady=(10, 0))

        filter_page = ttk.Frame(self.notebook, padding=8)
        parameter_page = ttk.Frame(self.notebook, padding=8)
        spectrum_page = ttk.Frame(self.notebook, padding=8)
        log_page = ttk.Frame(self.notebook, padding=8)
        self.notebook.add(filter_page, text="筛选下载")
        self.notebook.add(parameter_page, text="另存字段")
        self.notebook.add(spectrum_page, text="光谱可视化")
        self.notebook.add(log_page, text="日志")

        self._build_filter_page(filter_page)
        self._build_parameter_page(parameter_page)
        self._build_spectrum_page(spectrum_page)
        self._build_log_page(log_page)

        status_bar = ttk.Label(outer, textvariable=self.status_var, anchor="w")
        status_bar.pack(fill="x", pady=(8, 0))

    def _configure_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Title.TLabel", font=("Microsoft YaHei UI", 16, "bold"))
        style.configure("Subtle.TLabel", foreground="#555555")

    def _build_menu(self):
        menu_bar = tk.Menu(self)

        file_menu = tk.Menu(menu_bar, tearoff=False)
        file_menu.add_command(label="加载汇总表", command=self.load_summary_async)
        file_menu.add_command(label="下载选中数据", command=self.export_async)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.destroy)
        menu_bar.add_cascade(label="文件", menu=file_menu)

        feature_menu = tk.Menu(menu_bar, tearoff=False)
        feature_menu.add_command(label="筛选下载", command=lambda: self.notebook.select(0))
        feature_menu.add_command(label="另存字段", command=lambda: self.notebook.select(1))
        feature_menu.add_command(label="光谱可视化", command=lambda: self.notebook.select(2))
        menu_bar.add_cascade(label="功能", menu=feature_menu)

        help_menu = tk.Menu(menu_bar, tearoff=False)
        help_menu.add_command(label="关于", command=self.show_about)
        menu_bar.add_cascade(label="帮助", menu=help_menu)

        self.config(menu=menu_bar)

    def _path_row(self, parent, label, var, command, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(8, 8), pady=4)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", pady=4)
        ttk.Button(parent, text="选择", command=command, width=8).grid(row=row, column=2, padx=8, pady=4)

    def _build_filter_page(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=2)
        parent.columnconfigure(2, weight=2)
        parent.rowconfigure(1, weight=1)

        ttk.Label(parent, text="1. 选择字段").grid(row=0, column=0, sticky="w", padx=4)
        ttk.Label(parent, text="2. 选择该字段允许的值").grid(row=0, column=1, sticky="w", padx=4)
        ttk.Label(parent, text="3. 已启用筛选条件").grid(row=0, column=2, sticky="w", padx=4)

        field_frame = ttk.Frame(parent)
        field_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 8), pady=(4, 0))
        field_frame.rowconfigure(0, weight=1)
        field_frame.columnconfigure(0, weight=1)
        self.field_listbox = tk.Listbox(field_frame, exportselection=False)
        field_scroll = ttk.Scrollbar(field_frame, orient="vertical", command=self.field_listbox.yview)
        self.field_listbox.configure(yscrollcommand=field_scroll.set)
        self.field_listbox.grid(row=0, column=0, sticky="nsew")
        field_scroll.grid(row=0, column=1, sticky="ns")
        self.field_listbox.bind("<<ListboxSelect>>", self.on_field_selected)

        value_frame = ttk.Frame(parent)
        value_frame.grid(row=1, column=1, sticky="nsew", padx=8, pady=(4, 0))
        value_frame.rowconfigure(2, weight=1)
        value_frame.columnconfigure(0, weight=1)
        ttk.Entry(value_frame, textvariable=self.value_search_var).grid(row=0, column=0, columnspan=3, sticky="ew")
        self.value_search_var.trace_add("write", lambda *_: self.refresh_value_list())
        ttk.Button(value_frame, text="全选当前", command=self.select_all_visible_values).grid(row=1, column=0, sticky="w", pady=6)
        ttk.Button(value_frame, text="清空选择", command=self.clear_value_selection).grid(row=1, column=1, sticky="w", padx=6, pady=6)
        ttk.Button(value_frame, text="加入/更新条件", command=self.add_or_update_filter).grid(row=1, column=2, sticky="e", pady=6)
        self.value_listbox = tk.Listbox(value_frame, selectmode="extended", exportselection=False)
        value_scroll = ttk.Scrollbar(value_frame, orient="vertical", command=self.value_listbox.yview)
        self.value_listbox.configure(yscrollcommand=value_scroll.set)
        self.value_listbox.grid(row=2, column=0, columnspan=3, sticky="nsew")
        value_scroll.grid(row=2, column=3, sticky="ns")

        active_frame = ttk.Frame(parent)
        active_frame.grid(row=1, column=2, sticky="nsew", padx=(8, 0), pady=(4, 0))
        active_frame.rowconfigure(0, weight=1)
        active_frame.columnconfigure(0, weight=1)
        self.filter_tree = ttk.Treeview(active_frame, columns=("field", "count", "values"), show="headings", height=12)
        self.filter_tree.heading("field", text="字段")
        self.filter_tree.heading("count", text="值数量")
        self.filter_tree.heading("values", text="已选值")
        self.filter_tree.column("field", width=110, anchor="w")
        self.filter_tree.column("count", width=60, anchor="center")
        self.filter_tree.column("values", width=320, anchor="w")
        tree_scroll = ttk.Scrollbar(active_frame, orient="vertical", command=self.filter_tree.yview)
        self.filter_tree.configure(yscrollcommand=tree_scroll.set)
        self.filter_tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll.grid(row=0, column=1, sticky="ns")
        tree_buttons = ttk.Frame(active_frame)
        tree_buttons.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        ttk.Button(tree_buttons, text="删除选中条件", command=self.remove_selected_filter).pack(side="left")
        ttk.Button(tree_buttons, text="清空所有条件", command=self.clear_all_filters).pack(side="left", padx=8)

        tip = (
            "说明：同一字段内多选值表示“满足任意一个值”；多个字段之间由上方 AND/OR 控制。"
            "没有启用任何条件时，默认导出全部记录。"
        )
        ttk.Label(parent, text=tip, style="Subtle.TLabel").grid(row=2, column=0, columnspan=3, sticky="w", pady=(8, 0))

    def _build_parameter_page(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        self.parameter_box = MultiSelectBox(parent, "选择要另存到汇总表中的字段", height=24)
        self.parameter_box.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)

    def _build_spectrum_page(self, parent):
        parent.columnconfigure(0, weight=0)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(0, weight=1)

        control_frame = ttk.Frame(parent)
        control_frame.grid(row=0, column=0, sticky="ns", padx=(0, 10))
        control_frame.columnconfigure(0, weight=1)

        sample_frame = ttk.LabelFrame(control_frame, text="样本")
        sample_frame.grid(row=0, column=0, sticky="ew")
        sample_frame.columnconfigure(0, weight=1)
        self.spectrum_selector = ttk.Combobox(sample_frame, textvariable=self.spectrum_sample_var, values=[], width=30)
        self.spectrum_selector.grid(row=0, column=0, columnspan=2, sticky="ew", padx=8, pady=(8, 4))
        ttk.Button(sample_frame, text="绘制光谱", command=self.plot_selected_spectrum).grid(row=1, column=0, sticky="ew", padx=(8, 4), pady=4)
        ttk.Button(sample_frame, text="筛选首条", command=self.use_first_filtered_spectrum).grid(row=1, column=1, sticky="ew", padx=(4, 8), pady=4)
        ttk.Button(sample_frame, text="清空图像", command=self.clear_spectrum_plot).grid(row=2, column=0, columnspan=2, sticky="ew", padx=8, pady=(4, 8))

        process_frame = ttk.LabelFrame(control_frame, text="处理")
        process_frame.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        process_frame.columnconfigure(1, weight=1)
        ttk.Label(process_frame, text="起始波段").grid(row=0, column=0, sticky="w", padx=8, pady=(8, 4))
        ttk.Entry(process_frame, textvariable=self.spectrum_min_band_var, width=12).grid(row=0, column=1, sticky="ew", padx=8, pady=(8, 4))
        ttk.Label(process_frame, text="结束波段").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        ttk.Entry(process_frame, textvariable=self.spectrum_max_band_var, width=12).grid(row=1, column=1, sticky="ew", padx=8, pady=4)
        ttk.Label(process_frame, text="平滑窗口").grid(row=2, column=0, sticky="w", padx=8, pady=4)
        ttk.Spinbox(process_frame, from_=1, to=101, increment=2, textvariable=self.spectrum_smooth_var, width=10).grid(row=2, column=1, sticky="ew", padx=8, pady=4)
        ttk.Checkbutton(process_frame, text="0-1 归一化", variable=self.spectrum_normalize_var).grid(row=3, column=0, columnspan=2, sticky="w", padx=8, pady=4)
        ttk.Button(process_frame, text="重新应用处理", command=self.redraw_current_spectrum).grid(row=4, column=0, columnspan=2, sticky="ew", padx=8, pady=(4, 8))

        meta_frame = ttk.LabelFrame(control_frame, text="样本信息")
        meta_frame.grid(row=2, column=0, sticky="nsew", pady=(10, 0))
        control_frame.rowconfigure(2, weight=1)
        meta_frame.rowconfigure(0, weight=1)
        meta_frame.columnconfigure(0, weight=1)
        self.spectrum_meta_tree = ttk.Treeview(meta_frame, columns=("field", "value"), show="headings", height=12)
        self.spectrum_meta_tree.heading("field", text="字段")
        self.spectrum_meta_tree.heading("value", text="值")
        self.spectrum_meta_tree.column("field", width=96, anchor="w")
        self.spectrum_meta_tree.column("value", width=180, anchor="w")
        meta_scroll = ttk.Scrollbar(meta_frame, orient="vertical", command=self.spectrum_meta_tree.yview)
        self.spectrum_meta_tree.configure(yscrollcommand=meta_scroll.set)
        self.spectrum_meta_tree.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=8)
        meta_scroll.grid(row=0, column=1, sticky="ns", pady=8, padx=(0, 8))

        plot_frame = ttk.LabelFrame(parent, text="光谱曲线")
        plot_frame.grid(row=0, column=1, sticky="nsew")
        plot_frame.columnconfigure(0, weight=1)
        plot_frame.rowconfigure(0, weight=1)
        self.spectrum_canvas = tk.Canvas(plot_frame, bg="white", highlightthickness=1, highlightbackground="#cccccc")
        self.spectrum_canvas.grid(row=0, column=0, sticky="nsew", padx=8, pady=(8, 4))
        self.spectrum_canvas.bind("<Configure>", lambda _event: self.redraw_current_spectrum())
        ttk.Label(plot_frame, textvariable=self.spectrum_stats_var, style="Subtle.TLabel").grid(row=1, column=0, sticky="w", padx=8, pady=(0, 8))

    def _build_log_page(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        self.log_text = tk.Text(parent, wrap="word", height=12)
        log_scroll = ttk.Scrollbar(parent, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        log_scroll.grid(row=0, column=1, sticky="ns")

    def browse_spectra_dir(self):
        path = filedialog.askdirectory(title="选择光谱文件夹")
        if path:
            self.spectra_dir_var.set(path)

    def browse_summary_path(self):
        path = filedialog.askopenfilename(
            title="选择汇总表",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
        )
        if path:
            self.summary_path_var.set(path)

    def browse_output_dir(self):
        path = filedialog.askdirectory(title="选择导出位置")
        if path:
            self.output_dir_var.set(path)

    def show_about(self):
        messagebox.showinfo(
            "关于",
            f"{APP_TITLE}\n\n"
            "滞尘数据库检索、下载与光谱可视化工具。\n\n"
            f"第一维护者：{MAINTAINER}\n邮箱：{MAINTAINER_EMAIL}",
        )

    def log(self, message):
        timestamp = dt.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)

    def set_busy(self, busy, message=None):
        if message:
            self.status_var.set(message)
        state = "disabled" if busy else "normal"
        for child in self.winfo_children():
            self._set_state_recursive(child, state)

    def _set_state_recursive(self, widget, state):
        try:
            if isinstance(widget, (ttk.Button, ttk.Entry, ttk.Checkbutton, ttk.Combobox, ttk.Radiobutton)):
                widget.configure(state=state)
        except tk.TclError:
            pass
        for child in widget.winfo_children():
            self._set_state_recursive(child, state)

    def load_summary_async(self):
        thread = threading.Thread(target=self._load_summary_worker, daemon=True)
        thread.start()

    def _load_summary_worker(self):
        with self.load_lock:
            self.after(0, lambda: self.set_busy(True, "正在加载汇总表..."))
            try:
                summary_path = Path(self.summary_path_var.get().strip())
                rows, headers = self.load_summary(summary_path)
                self.rows = rows
                self.headers = headers
                self.filter_fields = headers[:]
                self.field_values = self.build_field_values(rows, headers)
                self.active_filters = {}
                self.current_field = ""
                self.after(0, self.populate_after_load)
                self.after(0, lambda: self.log(f"已加载 {len(rows)} 条记录，字段 {len(headers)} 个。"))
                self.after(0, lambda: self.count_var.set(f"已加载 {len(rows)} 条记录"))
                self.after(0, lambda: self.status_var.set("汇总表加载完成，可以添加筛选条件。"))
            except Exception as exc:
                detail = traceback.format_exc()
                self.after(0, lambda: self.log(detail))
                self.after(0, lambda: messagebox.showerror("加载失败", str(exc)))
                self.after(0, lambda: self.status_var.set("加载失败。"))
            finally:
                self.after(0, lambda: self.set_busy(False))

    def load_summary(self, path):
        if openpyxl is None:
            raise RuntimeError("缺少 openpyxl。请先运行 setup_env.ps1 安装最小环境。")
        if not str(path):
            raise RuntimeError("请先选择汇总表路径。")
        if not path.exists():
            raise FileNotFoundError(f"汇总表不存在：{path}")

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        all_rows = []
        common_headers = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(values_only=True, min_row=1, max_row=1), None)
            if not header_row:
                continue
            headers = [normalize_value(v) for v in header_row if normalize_value(v)]
            if common_headers is None:
                common_headers = ["工作表"] + headers
            for values in ws.iter_rows(values_only=True, min_row=2):
                if not any(normalize_value(v) for v in values):
                    continue
                record = {"工作表": sheet_name}
                for index, header in enumerate(headers):
                    record[header] = normalize_value(values[index] if index < len(values) else "")
                if record.get("标识码"):
                    all_rows.append(record)
        if not common_headers:
            raise RuntimeError("没有读取到汇总表表头。")
        if "标识码" not in common_headers:
            raise RuntimeError("汇总表缺少“标识码”字段。")
        return all_rows, common_headers

    def build_field_values(self, rows, fields):
        values = {}
        for field in fields:
            unique = {display_value(row.get(field, "")) for row in rows}
            values[field] = sorted(unique, key=lambda x: (x == BLANK_LABEL, x))
        return values

    def populate_after_load(self):
        self.field_listbox.delete(0, tk.END)
        for field in self.filter_fields:
            self.field_listbox.insert(tk.END, field)
        if self.filter_fields:
            self.field_listbox.selection_set(0)
            self.on_field_selected()

        self.parameter_box.set_values(self.headers, select_all=True)
        self.update_spectrum_selector()
        self.refresh_filter_tree()

    def on_field_selected(self, event=None):
        selection = self.field_listbox.curselection()
        if not selection:
            return
        self.current_field = self.field_listbox.get(selection[0])
        self.value_search_var.set("")
        self.refresh_value_list()

    def refresh_value_list(self):
        self.value_listbox.delete(0, tk.END)
        if not self.current_field:
            self.current_visible_values = []
            return
        query = self.value_search_var.get().strip().lower()
        values = self.field_values.get(self.current_field, [])
        if query:
            values = [v for v in values if query in v.lower()]
        self.current_visible_values = values
        selected_values = self.active_filters.get(self.current_field, set())
        for value in values:
            self.value_listbox.insert(tk.END, value)
        for index, value in enumerate(values):
            if stored_value(value) in selected_values:
                self.value_listbox.selection_set(index)
        self.status_var.set(f"字段“{self.current_field}”可选值 {len(values)} 个。")

    def select_all_visible_values(self):
        self.value_listbox.selection_set(0, tk.END)

    def clear_value_selection(self):
        self.value_listbox.selection_clear(0, tk.END)

    def add_or_update_filter(self):
        if not self.current_field:
            messagebox.showwarning("未选择字段", "请先选择一个字段。")
            return
        selected = [self.value_listbox.get(i) for i in self.value_listbox.curselection()]
        if not selected:
            messagebox.showwarning("未选择字段值", "请至少选择一个字段值。")
            return
        self.active_filters[self.current_field] = {stored_value(v) for v in selected}
        self.refresh_filter_tree()
        self.log(f"已更新条件：{self.current_field} = {', '.join(selected[:8])}{' ...' if len(selected) > 8 else ''}")

    def remove_selected_filter(self):
        selection = self.filter_tree.selection()
        if not selection:
            return
        for item in selection:
            field = self.filter_tree.item(item, "values")[0]
            self.active_filters.pop(field, None)
        self.refresh_filter_tree()
        self.refresh_value_list()

    def clear_all_filters(self):
        self.active_filters = {}
        self.refresh_filter_tree()
        self.refresh_value_list()
        self.log("已清空所有筛选条件。")

    def refresh_filter_tree(self):
        for item in self.filter_tree.get_children():
            self.filter_tree.delete(item)
        for field, values in self.active_filters.items():
            display_values = [display_value(v) for v in sorted(values)]
            text = ", ".join(display_values[:8])
            if len(display_values) > 8:
                text += " ..."
            self.filter_tree.insert("", tk.END, values=(field, len(values), text))

    def selected_parameter_fields(self):
        selected = self.parameter_box.get_selected()
        fields = []
        for field in ["工作表", "标识码"]:
            if field in self.headers and field not in fields:
                fields.append(field)
        for field in selected:
            if field in self.headers and field not in fields:
                fields.append(field)
        return fields

    def update_spectrum_selector(self):
        sample_ids = [row.get("标识码", "") for row in self.rows if row.get("标识码", "")]
        if self.spectrum_selector is not None:
            self.spectrum_selector.configure(values=sample_ids)
        if sample_ids and self.spectrum_sample_var.get() not in sample_ids:
            self.spectrum_sample_var.set(sample_ids[0])

    def sample_row_by_id(self, sample_id):
        for row in self.rows:
            if row.get("标识码", "") == sample_id:
                return row
        return None

    def populate_spectrum_metadata(self, row):
        if self.spectrum_meta_tree is None:
            return
        for item in self.spectrum_meta_tree.get_children():
            self.spectrum_meta_tree.delete(item)
        if not row:
            return
        priority_fields = [
            "标识码",
            "树种",
            "植被类型",
            "采样日期",
            "季度",
            "采样地点",
            "光谱尺度",
            "传感器",
            "滞尘能力（g/m²）",
        ]
        fields = [field for field in priority_fields if field in row]
        fields.extend(field for field in self.headers if field not in fields and field in row)
        for field in fields:
            value = row.get(field, "")
            if value != "":
                self.spectrum_meta_tree.insert("", tk.END, values=(field, value))

    def use_first_filtered_spectrum(self):
        try:
            for row in self.filter_rows():
                sample_id = row.get("标识码", "")
                if self.spectrum_exists(sample_id):
                    self.spectrum_sample_var.set(sample_id)
                    self.plot_selected_spectrum()
                    return
            messagebox.showwarning("没有可绘制光谱", "当前筛选结果中没有找到匹配的光谱 txt。")
        except Exception as exc:
            messagebox.showerror("选择失败", str(exc))

    def plot_selected_spectrum(self):
        sample_id = self.spectrum_sample_var.get().strip()
        if not sample_id:
            messagebox.showwarning("未选择样本", "请先输入或选择一个标识码。")
            return
        try:
            points = self.load_spectrum(sample_id)
            self.current_spectrum = (sample_id, points)
            self.populate_spectrum_metadata(self.sample_row_by_id(sample_id))
            self.redraw_current_spectrum()
            self.notebook.select(2)
            self.log(f"已绘制光谱：{sample_id}，波段点 {len(points)} 个。")
        except Exception as exc:
            self.current_spectrum = None
            self.clear_spectrum_plot()
            messagebox.showerror("绘制失败", str(exc))

    def load_spectrum(self, sample_id):
        path = self.spectrum_path(sample_id)
        if not self.spectra_dir_var.get().strip():
            raise RuntimeError("请先选择光谱文件夹。")
        if not path.exists():
            raise FileNotFoundError(f"光谱 txt 不存在：{path}")
        points = []
        text = path.read_text(encoding="utf-8-sig")
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            if "," in line:
                parts = [part.strip() for part in line.split(",")]
            elif "\t" in line:
                parts = [part.strip() for part in line.split("\t")]
            else:
                parts = line.split()
            if len(parts) < 2:
                continue
            try:
                band = float(parts[0])
                reflectance = float(parts[1])
            except ValueError:
                continue
            if math.isfinite(band) and math.isfinite(reflectance):
                points.append((band, reflectance))
        if not points:
            raise RuntimeError(f"未从光谱文件读取到有效数值：{path}")
        points.sort(key=lambda item: item[0])
        return points

    def processed_spectrum(self, points):
        min_band = self._optional_float(self.spectrum_min_band_var.get())
        max_band = self._optional_float(self.spectrum_max_band_var.get())
        filtered = []
        for band, reflectance in points:
            if min_band is not None and band < min_band:
                continue
            if max_band is not None and band > max_band:
                continue
            filtered.append((band, reflectance))
        if not filtered:
            raise RuntimeError("当前波段范围内没有数据点。")

        window = self._smooth_window()
        if window > 1 and len(filtered) >= window:
            half = window // 2
            smoothed = []
            values = [value for _band, value in filtered]
            for index, (band, _value) in enumerate(filtered):
                start = max(0, index - half)
                end = min(len(values), index + half + 1)
                smoothed.append((band, sum(values[start:end]) / (end - start)))
            filtered = smoothed

        if self.spectrum_normalize_var.get():
            values = [value for _band, value in filtered]
            low = min(values)
            high = max(values)
            if high > low:
                filtered = [(band, (value - low) / (high - low)) for band, value in filtered]
            else:
                filtered = [(band, 0.0) for band, _value in filtered]
        return filtered

    def _optional_float(self, value):
        value = value.strip()
        if not value:
            return None
        return float(value)

    def _smooth_window(self):
        try:
            window = int(self.spectrum_smooth_var.get())
        except (tk.TclError, ValueError):
            return 1
        if window < 1:
            return 1
        if window % 2 == 0:
            window += 1
            self.spectrum_smooth_var.set(window)
        return window

    def redraw_current_spectrum(self):
        if self.spectrum_canvas is None:
            return
        if not self.current_spectrum:
            self.draw_blank_spectrum()
            return
        sample_id, points = self.current_spectrum
        try:
            processed = self.processed_spectrum(points)
        except Exception as exc:
            self.draw_blank_spectrum(str(exc))
            return
        self.draw_spectrum(sample_id, processed)

    def clear_spectrum_plot(self):
        self.current_spectrum = None
        self.populate_spectrum_metadata(None)
        self.draw_blank_spectrum()

    def draw_blank_spectrum(self, message="选择样本后绘制光谱"):
        canvas = self.spectrum_canvas
        if canvas is None:
            return
        canvas.delete("all")
        width = max(canvas.winfo_width(), 400)
        height = max(canvas.winfo_height(), 280)
        canvas.create_text(width / 2, height / 2, text=message, fill="#666666", font=("Microsoft YaHei UI", 12))
        self.spectrum_stats_var.set("未绘制光谱")

    def draw_spectrum(self, sample_id, points):
        canvas = self.spectrum_canvas
        canvas.delete("all")
        width = max(canvas.winfo_width(), 520)
        height = max(canvas.winfo_height(), 360)
        left, right, top, bottom = 64, 20, 28, 48
        plot_width = max(width - left - right, 1)
        plot_height = max(height - top - bottom, 1)

        bands = [band for band, _value in points]
        values = [value for _band, value in points]
        min_x, max_x = min(bands), max(bands)
        min_y, max_y = min(values), max(values)
        if max_x == min_x:
            max_x = min_x + 1
        if max_y == min_y:
            max_y = min_y + 1
        y_pad = (max_y - min_y) * 0.05
        min_y -= y_pad
        max_y += y_pad

        def sx(x_value):
            return left + (x_value - min_x) / (max_x - min_x) * plot_width

        def sy(y_value):
            return top + (max_y - y_value) / (max_y - min_y) * plot_height

        canvas.create_rectangle(left, top, left + plot_width, top + plot_height, outline="#999999")
        for index in range(6):
            x = left + plot_width * index / 5
            y = top + plot_height * index / 5
            canvas.create_line(x, top, x, top + plot_height, fill="#eeeeee")
            canvas.create_line(left, y, left + plot_width, y, fill="#eeeeee")
            x_value = min_x + (max_x - min_x) * index / 5
            y_value = max_y - (max_y - min_y) * index / 5
            canvas.create_text(x, top + plot_height + 16, text=f"{x_value:.0f}", fill="#555555", font=("Microsoft YaHei UI", 8))
            canvas.create_text(left - 8, y, text=f"{y_value:.3f}", anchor="e", fill="#555555", font=("Microsoft YaHei UI", 8))

        line_points = []
        for band, value in points:
            line_points.extend([sx(band), sy(value)])
        if len(line_points) >= 4:
            canvas.create_line(*line_points, fill="#1f6feb", width=2, smooth=False)

        canvas.create_text(left + plot_width / 2, height - 14, text="Bands (nm)", fill="#333333", font=("Microsoft YaHei UI", 9))
        canvas.create_text(18, top + plot_height / 2, text="Reflectance", fill="#333333", angle=90, font=("Microsoft YaHei UI", 9))
        canvas.create_text(left, 14, text=sample_id, anchor="w", fill="#222222", font=("Microsoft YaHei UI", 10, "bold"))

        mean_y = sum(values) / len(values)
        self.spectrum_stats_var.set(
            f"点数 {len(points)}；波段 {min_x:.0f}-{max_x:.0f} nm；"
            f"反射率 min {min(values):.4f} / mean {mean_y:.4f} / max {max(values):.4f}"
        )

    def spectrum_path(self, sample_id):
        spectra_dir = Path(self.spectra_dir_var.get().strip())
        return spectra_dir / f"{sample_id}.txt"

    def spectrum_exists(self, sample_id):
        spectra_dir = self.spectra_dir_var.get().strip()
        if not spectra_dir:
            return False
        return self.spectrum_path(sample_id).exists()

    def row_matches_filter(self, row):
        if not self.active_filters:
            return True
        checks = []
        for field, allowed in self.active_filters.items():
            checks.append(row.get(field, "") in allowed)
        if self.filter_logic_var.get() == "OR":
            return any(checks)
        return all(checks)

    def filter_rows(self):
        if not self.rows:
            raise RuntimeError("请先加载汇总表。")
        return [row for row in self.rows if self.row_matches_filter(row)]

    def preview_count(self):
        try:
            rows = self.filter_rows()
            if self.spectra_dir_var.get().strip():
                existing = sum(1 for row in rows if self.spectrum_exists(row.get("标识码", "")))
                missing = len(rows) - existing
                message = f"匹配 {len(rows)} 条；有 txt {existing} 条；缺失 {missing} 条"
            else:
                message = f"匹配 {len(rows)} 条；未选择光谱文件夹，暂不检查 txt"
            self.count_var.set(message)
            self.log(f"预览：{message}。")
        except Exception as exc:
            messagebox.showerror("预览失败", str(exc))

    def export_async(self):
        thread = threading.Thread(target=self._export_worker, daemon=True)
        thread.start()

    def _export_worker(self):
        self.after(0, lambda: self.set_busy(True, "正在导出..."))
        try:
            rows = self.filter_rows()
            if not rows:
                raise RuntimeError("当前筛选条件没有匹配记录。")
            if not self.output_dir_var.get().strip():
                raise RuntimeError("请先选择导出位置。")
            if self.copy_spectra_var.get() and not self.spectra_dir_var.get().strip():
                raise RuntimeError("已勾选下载光谱 txt，请先选择光谱文件夹。")

            fields = self.selected_parameter_fields()
            if not fields:
                raise RuntimeError("请至少选择一个另存字段。")

            output_root = Path(self.output_dir_var.get().strip())
            export_dir = output_root / safe_filename(f"{APP_NAME}_{APP_VERSION}_export_" + dt.datetime.now().strftime("%Y%m%d_%H%M%S"))
            export_dir.mkdir(parents=True, exist_ok=False)

            copied = 0
            missing = []
            if self.copy_spectra_var.get():
                spectra_output = export_dir / "光谱"
                spectra_output.mkdir(parents=True, exist_ok=True)
                for index, row in enumerate(rows, start=1):
                    sample_id = row.get("标识码", "")
                    source = self.spectrum_path(sample_id)
                    if source.exists():
                        shutil.copy2(source, spectra_output / source.name)
                        copied += 1
                    else:
                        missing.append(row)
                    if index % 500 == 0:
                        self.after(0, lambda i=index: self.status_var.set(f"正在复制光谱：{i}/{len(rows)}"))
            elif self.spectra_dir_var.get().strip():
                for row in rows:
                    sample_id = row.get("标识码", "")
                    if not self.spectrum_exists(sample_id):
                        missing.append(row)

            if self.save_csv_var.get():
                self.write_csv(export_dir / "筛选汇总表.csv", rows, fields)
            if self.save_xlsx_var.get():
                self.write_xlsx(export_dir / "筛选汇总表.xlsx", rows, fields)
            self.write_csv(export_dir / "缺失光谱清单.csv", missing, fields)
            self.write_log(export_dir / "导出日志.txt", rows, fields, copied, missing)

            message = f"导出完成：匹配 {len(rows)} 条，复制 txt {copied} 个，缺失 {len(missing)} 个。"
            self.after(0, lambda: self.log(message))
            self.after(0, lambda: self.status_var.set(message))
            self.after(0, lambda: self.count_var.set(message))
            self.after(0, lambda: messagebox.showinfo("导出完成", f"{message}\n\n输出目录：\n{export_dir}"))
        except Exception as exc:
            detail = traceback.format_exc()
            self.after(0, lambda: self.log(detail))
            self.after(0, lambda: messagebox.showerror("导出失败", str(exc)))
            self.after(0, lambda: self.status_var.set("导出失败。"))
        finally:
            self.after(0, lambda: self.set_busy(False))

    def write_csv(self, path, rows, fields):
        with path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({field: row.get(field, "") for field in fields})

    def write_xlsx(self, path, rows, fields):
        if Workbook is None:
            raise RuntimeError("缺少 openpyxl，无法另存 XLSX。")
        wb = Workbook()
        ws = wb.active
        ws.title = "筛选结果"
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_length + 2, 10), 32)
        wb.save(path)

    def write_log(self, path, rows, fields, copied, missing):
        lines = [
            f"软件：{APP_TITLE}",
            f"第一维护者：{MAINTAINER}",
            f"邮箱：{MAINTAINER_EMAIL}",
            f"导出时间：{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"光谱文件夹：{self.spectra_dir_var.get().strip()}",
            f"汇总表路径：{self.summary_path_var.get().strip()}",
            f"多字段关系：{self.filter_logic_var.get()}",
            f"导出记录数：{len(rows)}",
            f"复制光谱数：{copied}",
            f"缺失光谱数：{len(missing)}",
            f"另存字段：{', '.join(fields)}",
            "",
            "筛选条件：",
        ]
        if not self.active_filters:
            lines.append("无，导出全部记录。")
        else:
            for field, values in self.active_filters.items():
                values_text = ", ".join(display_value(v) for v in sorted(values))
                lines.append(f"{field}: {values_text}")
        path.write_text("\n".join(lines), encoding="utf-8")


def main():
    if openpyxl is None:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("缺少依赖", "缺少 openpyxl。请先在工具目录运行 setup_env.ps1。")
        return 1
    app = DustDatabaseDownloader()
    app.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())
